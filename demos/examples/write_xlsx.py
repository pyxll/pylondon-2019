from openpyxl import Workbook

wb = Workbook()

# Get the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Set style
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
ws['A1'].font = Font(color=RED)

# Rows can also be appended
ws.append([1, 2, 3])

# Python types converted
import datetime
ws['A2'] = datetime.datetime.now()

# Write a pandas DataFrame
import pandas as pd
import os

filename = os.path.join(os.path.dirname(__file__), "..", "data", "zoo.csv")
df = pd.read_csv(filename, parse_dates=True, infer_datetime_format=True)

from openpyxl.utils.dataframe import dataframe_to_rows

ws.append([])
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Save the file
wb.save("xlsx_example.xlsx")
print("Written xlsx file")
